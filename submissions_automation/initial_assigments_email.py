#!/bin/usr/env python3

import os
import openpyxl
import email_assistant
from config import *
import time
import datetime
import shutil

wb = openpyxl.load_workbook(filename = SUBMITTERS_PATH)
ws = wb['Submitters']

for row in range(2, ws.max_row+1):
    submitter_id = ws.cell(row = row, column = 1).value
    submitter_first = ws.cell(row = row, column = 2).value
    submitter_last = ws.cell(row = row, column = 3).value
    submitter_email = ws.cell(row = row, column = 4).value
    #submitter_email = 'david@reflashed.com'
    
    submitter_ws = wb['{}'.format(submitter_id)]
    
    # Sloppy...
    do_break = False
    # Send out their new assignement
    for subdir, dirs, files in os.walk(UNASSIGNED_ASSIGNMENTS_DIR):
        if do_break == True:
            break
        
        for file in files:
            file_path = os.path.join(subdir, file)
            extension = os.path.splitext(os.path.basename(file_path))[1]
            
            if extension == '.xlsx':
                new_assignment_id = os.path.splitext(os.path.basename(file_path))[0]
                do_break = True
                break
            
    # Create the assignment
    if(new_assignment_id):
        save_row = submitter_ws.max_row+1
        
        # Move the old assignment
        shutil.move('{}{}\\'.format(UNASSIGNED_ASSIGNMENTS_DIR, new_assignment_id), '{}{}\\'.format(ASSIGNED_ASSIGNMENTS_DIR, new_assignment_id))
                    
        # Attach everything in this dir
        attachment_dir = '{}{}'.format(ASSIGNED_ASSIGNMENTS_DIR, new_assignment_id)
        attachments = ['{}{}\\{}'.format(ASSIGNED_ASSIGNMENTS_DIR, new_assignment_id, file) for file in os.listdir(attachment_dir)]

        asst = email_assistant.EmailAssistant(USERNAME, PASSWORD)
        asst.send_email(submitter_email, 'Mason Lab Microbe Database - New Assignment', 'Thank you for your previous submission! This set of annotations corrresponds to assignment ID {}. Recall, your submitter ID is {}. Whenever you submit this assignment, the subject line must read {}_{}. This is extremely important!! Please do not reply to this email with anything but submissions! Send all questions to hebashaaban1@gmail.com and daw2035@med.cornell.edu. Keep up the good work!'.format(new_assignment_id, submitter_id, submitter_id, new_assignment_id), attachments)
        
        # Log that we sent out a new assignment
        submitter_ws.cell(row = save_row, column = 1).value = new_assignment_id
        
        # Timestamp
        ts = time.time()
        timestamp = datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')
        submitter_ws.cell(row = save_row, column = 2).value = timestamp
        wb.save(SUBMITTERS_PATH)
        
    