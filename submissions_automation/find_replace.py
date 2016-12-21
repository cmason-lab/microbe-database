#!/usr/bin/env python3
 # coding=utf-8

import openpyxl
import re

path = 'annotations_master 12-19-16 - Copy.xlsx'

keys_to_cols = {'optimal_pH': 13, 'optimal_temperature': 15, 'pathogenicity': 19,
                'is_susceptible_to_antibiotics': 29, 'spore_forming': 49, 'biofilm_forming': 51,
                'is_extremophile': 3, 'gram_stain': 41, 'found_in_microbiome': 7, 'plant_pathogen': 33,
                'animal_pathogen': 35}

type_from_key = {'optimal_pH': 'range', 'optimal_temperature': 'range', 'pathogenicity': '1-4',
                'is_susceptible_to_antibiotics': 'binary', 'spore_forming': 'binary',
                'biofilm_forming': 'binary', 'is_extremophile': 'binary', 'gram_stain': '0-2',
                'found_in_microbiome': 'binary', 'plant_pathogen': 'binary', 'animal_pathogen': 'binary'}

wb = openpyxl.load_workbook(filename = path)
ws = wb.get_active_sheet()

for row in range(2, ws.max_row+1):
    for key in keys_to_cols:
        col = keys_to_cols[key]
        value = str(ws.cell(row = row, column = col).value)
        citation = str(ws.cell(row = row, column = col+1).value)
        
        if value == '#N/A' or value == 'N/A' or value == 'None':
            continue
        
        # No citation? Do not pass go, do not collect $200
        if citation == '#N/A' or citation == 'N/A' or citation == '' or citation == 'None':
            ws.cell(row = row, column = col).value = ''
            continue
        
        type = type_from_key[key] 
        valid_type = True
        
        if type == 'range':  
            if key == 'optimal_temperature':
                # Get rid of '°' and 'C'
                value = value.replace('°', '')
                value = value.replace('C', '')
            
            # Convert 'to' to ':'
            value = value.replace('to', ':')
            
            # Match this pattern and reduce it to 'number:number'
            pattern = re.compile('^\s*(-?[0-9]+\.?[0-9]*)\s*[-:–]\s*(-?[0-9]+\.?[0-9]*)\s*$')
            # We also accept straight up numbers
            pattern2 = re.compile('^\s*(-?[0-9]+\.?[0-9]*)\s*$')
            
            if pattern.match(value):
                value = re.sub(pattern, '\\1:\\2', value)
            else:
                if pattern2.match(value):
                    value = re.sub(pattern2, '\\1', value)
                    valid_type = True
                else:
                    pattern3 = re.compile('^((-?[0-9]+\.?[0-9]*)([:](-?[0-9]+\.?[0-9]*))?)(,((-?[0-9]+\.?[0-9]*)([:](-?[0-9]+\.?[0-9]*))?))*$')
                    if pattern3.match(value) is False:
                        valid_type = False
        elif type == 'binary':
            pattern = re.compile('^\s*([01])\s*$')
            
            value = value.lower()
            value = value.replace('yes', '1')
            value = value.replace('no', '0')
            
            if pattern.match(value):
                # just trim whitespace
                value = re.sub(pattern, '\\1', value)
            else:
                valid_type = False
        elif type == '1-4':
            pattern = re.compile('^\s*([1-4])\s*$')
            
            if pattern.match(value):
                # just trim whitespace
                value = re.sub(pattern, '\\1', value)
            else:
                valid_type = False
        elif type == '0-2':
            pattern = re.compile('^\s*([0-2])\s*$')
            
            if pattern.match(value):
                # just trim whitespace
                value = re.sub(pattern, '\\1', value)
            else:
                valid_type = False
        
        if valid_type is False:
            value = '{} badformat {}'.format(value, key)
        
        # Update the ws with the new value
        ws.cell(row = row, column = col).value = value
    
wb.save(path)