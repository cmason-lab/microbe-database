#!/usr/bin/env python3

from config import *
import openpyxl
import os

import time
import datetime
ts = time.time()
timestamp = datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')
print(timestamp)

exit()

attachment_dir = '{}{}'.format(UNASSIGNED_ASSIGNMENTS_DIR, '1')
attachments = ['{}{}'.format(UNASSIGNED_ASSIGNMENTS_DIR, file) for file in os.listdir(attachment_dir)]

print(attachments)
exit()

wb = openpyxl.load_workbook(filename = SUBMITTERS_PATH)
submitter_ws = wb['1']

# Create a reverse lookup dictionary
write_row_from_annotation_id = {}

for row in range(2, submitter_ws.max_row+1):
    cell = submitter_ws.cell(row = row, column = 1)
    write_row_from_annotation_id[cell.value] = row
    
print(write_row_from_annotation_id[4])
#print(write_row_from_annotation_id['1'])

exit()

submitters_ws.cell(row = submitters_ws.max_row+1, column = 1).value = ''

for row in range(2, submitters_ws.max_row+1):
    submitter_id = submitters_ws.cell(row = row, column = 1).value
    print(submitter_id)

#wb.save(SUBMITTERS_PATH)