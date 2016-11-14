import urllib.parse
import urllib.request
import urllib.error
import ssl
import re
import webbrowser

import openpyxl
from numpy.distutils.fcompiler import none
import os

class AnnotationsAutomator():

    def __init__(self, species_file):
        ''' Set all the constants '''
        
        self.species_file = species_file
        
        # Not all the SSL's are verified
        context = ssl.create_default_context()
        context.check_hostname = False
        context.verify_mode = ssl.CERT_NONE
        ssl_handler = urllib.request.HTTPSHandler(0, context, False)
        
        # Create "opener" (OpenerDirector instance)
        opener = urllib.request.build_opener(ssl_handler)
        
        # Install the opener, so we can make normal calls to urlopen
        urllib.request.install_opener(opener)

        # And we're live
        self.load_species(species_file);
        
        # Results are formatted {'Database': html list}
        self.results = {}

    def load_species(self, filename):
        ''' Loads the species from a .csv (one species per line) into a list '''

        #self.organisms = ['Hyphomicrobium zavarzinii', 'Rhodomicrobium vannielii']
        #return
        self.organisms = []
        
        self.read_wb = openpyxl.load_workbook(filename = filename)
        self.read_ws = self.read_wb.get_active_sheet()
        
        for row in range(2, self.read_ws.max_row+1):
            cell = self.read_ws.cell(row = row, column = 2)
            
            if cell.value is None:
                break
            
            self.organisms.append(cell.value)

    def save_results(self):
        # We are going to save this as just the name of the file
        save_dir = '{}\\'.format(os.path.dirname(self.species_file))
        save_name = os.path.splitext(os.path.basename(self.species_file))[0]
        
        results = ''
        results += '<html>'
        results += '<head><title>Search for {}</title></head>'.format(save_name)
        results += '<body>'
        
        results += '<p><h1>Search for {}</h1></p>'.format(save_name)
        
        for key, value in self.results.items():
            results += '<p>'
            results += '<h2>{}</h2>'.format(key)
            results += '</p>'
            
            results += '<p>'
            results += value
            results += '</p>'
            
        results += '</body>'
        results += '</html>'

        file = open('{}{}_presearch.html'.format(save_dir, save_name), 'w')
        file.write(results)
        file.close
    
    def search_db(self, name, search_url, search_key, regex, base_url, post=False, optional_params={}):
        results = ''
        results += '<ol>'

        i = 0
        imax = len(self.organisms)
        
        for organism in self.organisms:
            i += 1
                
            print('Searching {} for {} ({} of {})'.format(name, organism, i, imax))

            if not post:
                searchEncoded = urllib.parse.urlencode({search_key: organism})
                url = '{}{}'.format(search_url, searchEncoded)
                req = urllib.request.Request(url)
                print("URL: " + url)
            else:
                searchEncoded = urllib.parse.urlencode({search_key: organism})

                if(optional_params):
                    searchEncoded = searchEncoded + "&" + urllib.parse.urlencode(optional_params)
                    
                req = urllib.request.Request(search_url, searchEncoded.encode('ASCII'))
                print("URL: " + search_url)

            try:
                with urllib.request.urlopen(req) as resp:
                    respHTML = resp.read()
                    #print("Response: " + respHTML.decode('UTF-8'))

                    genus = re.match('(.+) .+', organism).group(1)
                    print("Genus: " + genus)

                    species = re.match('.+ (.+)', organism).group(1)
                    print("Species: " + species)

                    # organism is the binomial: Genus species
                    search_str = regex.format(organism=organism, species=species, genus=genus)
                    print("RegEx: " + search_str)
                    m = re.search(search_str.encode('utf-8'), respHTML, re.IGNORECASE)
                    if(m):
                        url = base_url + m.group(1).decode('UTF-8')
                        results += '<li>'
                        results += '<a href="{}">{}</a>'.format(url, organism)
                        results += '</li>'
            except urllib.error.URLError:
                print('URL doesn\'t exist, skipiping')
            finally:
                print('Error')
                    
        results += '</ol>'
        
        self.results[name] = results
        
    def search_MicrobeWiki(self):
        self.search_db('MicrobeWiki', 'https://microbewiki.kenyon.edu/index.php?title=Special%3ASearch&profile=default&fulltext=Search&', 'search', '\<a href=\"(.+)\" title=\"{organism}\">{organism}\<\/a\>', 'https://microbewiki.kenyon.edu')

    def search_BacMap(self):
        self.search_db('BacMap', 'http://bacmap.wishartlab.com/genomes?utf8=%E2%9C%93&type=taxonomy&', 'query', '\<a href=\"(.*)\" title\=\"{organism}.*', 'http://bacmap.wishartlab.com')

    def search_ATCC(self):
        self.search_db('ATCC', 'https://www.atcc.org/Search_Results.aspx?dsNav=Ntk:PrimarySearch%7cgdas%7c3%7c,Ny:True,Ro:0,N:1000552&', 'searchTerms', 'href=\"(.+)\"\>\<I\>{organism}\<\/I\>', 'https://www.atcc.org')

    def search_VFDB(self):
        self.search_db('VFDB', 'http://www.mgc.ac.cn/cgi-bin/VFs/search.cgi', 'Keywords', 'href=\"(.+)\"\>{genus}\<\/a\>', 'http://www.mgc.ac.cn', True, {'Field': 'Genus'})

    def search_PATRIC(self):
        self.search_db('PATRIC', 'https://www.patricbrc.org/portal/portal/patric/GlobalSearch/GlobalSearchWindow?action=b&cacheability=PAGE&need=search&_dc=1470968370608&', 'keyword', '\[{{\"taxon_id\"\:(.+)\,\"taxon\_name\"\:\"{organism}\"\,\"taxon\_rank\"\:\"organism\"', 'https://www.patricbrc.org/portal/portal/patric/Taxon?cType=taxon&cId=')

    def search_ARDB(self):
        self.search_db('ARDB', 'http://ardb.cbcb.umd.edu/cgi/search.cgi?db=S&field=af&', 'term', '<a href="([^<]+)">{organism}</a>', 'http://ardb.cbcb.umd.edu')
        
    def search_GOLD(self):
        url = 'https://gold.jgi.doe.gov/organisms?setColumns=yes&Organism.GOLD+Organism+ID_options=equals&Organism.Organism+Name_options=equals&Organism.Genus_options=equals&Organism.Genus+Synonyms_options=equals&Organism.Species_options=equals&Organism.Species+Synonyms_options=equals&Organism.Subspecies_options=equals&Organism.Strain_options=equals&Organism.Culture+Collection_options=equals&Organism.Domain_options=equals&Organism.Phylum_options=equals&Organism.NCBI+Taxonomy+ID_options=equals&Organism.NCBI+Kingdom_options=equals&Organism.NCBI+Phylum_options=equals&Organism.NCBI+Class_options=equals&Organism.NCBI+Order_options=equals&Organism.NCBI+Family_options=equals&Organism.NCBI+Genus_options=equals&Organism.NCBI+Species_options=equals&Organism.Submitter%27s+Organism+Name_options=equals&Organism.Biosafety+Level_options=equals&Organism.Comments_options=equals&Organism.Annotator+Comments_options=equals&Organism.Cultured_options=equals&Organism.Culture+Type_options=equals&Organism.Organism+Type_options=equals&Organism.Uncultured+Type_options=equals&Organism.Common+Name_options=equals&Organism.Taxon+DOI_options=equals&Organism.Exemplar+DOI_options=equals&Organism.Exemplar+Name_options=equals&Organism.Genbank+16S+ID_options=equals&Organism.Serovar_options=equals&Organism.Strain+Info+ID_options=equals&Organism.Type+Strain_options=equals&Organism.Is+Public_options=equals&Organism.Cell+Diameter_options=equals&Organism.Cell+Shape_options=equals&Organism.Color_options=equals&Organism.Gram+Stain_options=equals&Organism.Motility_options=equals&Organism.Oxygen+Requirement_options=equals&Organism.Ph_options=equals&Organism.Salinity_options=equals&Organism.Pressure_options=equals&Organism.Sporulation_options=equals&Organism.Carbon+Source_options=equals&Organism.Symbiont+Name_options=equals&Organism.Symbiont+Taxon+ID_options=equals&Organism.Symbiotic+Physical+Interaction_options=equals&Organism.Symbiotic+Relationship_options=equals&Organism.Temperature+Optimum_options=equals&Organism.Temperature+Range_options=equals&Organism.Cell+Length_options=equals&Organism.Commercial+Strain_options=equals&Organism.Commercial+Strain+Comments_options=equals&Organism.Viral+Group_options=equals&Organism.Viral+Subgroup_options=equals&Organism.Sample+Collection+Site_options=equals&Organism.Sample+Isolation+Comments_options=equals&Organism.Sample+Collection+Method_options=equals&Organism.Cruise+Line+Name_options=equals&Organism.Contact+Name_options=equals&Organism.Contact+Email_options=equals&Organism.Isolation+Host+Name_options=equals&Organism.Host+Taxonomy+ID_options=equals&Organism.Host+Gender_options=equals&Organism.Host+Race_options=equals&Organism.Host+Age_options=equals&Organism.Host+Health+Condition_options=equals&Organism.Patient+Visit+Number_options=equals&Organism.Host+Medication_options=equals&Organism.Medical+Record+Number_options=equals&Organism.Host+Body+Site_options=equals&Organism.Host+Body+Subsite_options=equals&Organism.Host+Body+Product_options=equals&Organism.Host+Specificity+or+Range_options=equals&Organism.Host+Comments_options=equals&Organism.Ecosystem_options=equals&Organism.Ecosystem+Category_options=equals&Organism.Ecosystem+Type_options=equals&Organism.Ecosystem+Subtype_options=equals&Organism.Specific+Ecosystem_options=equals&Organism.Ecosystem+Suggestion_options=equals&Organism.Geographic+Location_options=equals&Organism.Latitude_options=equals&Organism.Longitude_options=equals&Organism.Longhurst+Code_options=equals&Organism.Altitude_options=equals&Organism.Depth_options=equals&Organism.Sample+Collection+Temperature_options=equals&Organism.Chlorophyll+Concentration_options=equals&Organism.Nitrate+Concentration_options=equals&Organism.Oxygen+Concentration_options=equals&Organism.Salinity+Concentration_options=equals&Organism.Growth+Conditions_options=equals&'
        self.search_db('GOLD', url, 'Organism.Species', '<a href="\?(.+)" class="blackLink">.+</a>', 'https://gold.jgi.doe.gov/organisms?')
        
    def search_HOMD(self):
        self.search_db('HOMD', 'http://www.homd.org/index.php?name=HOMD&file=index', 'word_search', '<td align="center" class=\'HOMDsubtitle3\'>([^<]+)</td><td align="left" class=\'HOMDsubtitle3\'><i>{genus}</td><td align="left" class=\'HOMDsubtitle3\'><i>{species}</i>', 'http://www.homd.org/index.php?name=HOMD&view=dynamic&oraltaxonid=', True)

    def search_BEI(self):
        regex = '<a id=".+" template="bacteria" atccnum=".+" href="(.+)" target="_self">.+{organism}</I>'
        self.search_db('BEI', 'https://www.beiresources.org/Catalog.aspx?f_instockflag=In+Stock%23~%23Discontinued%23~%23Temporarily+Out+of+Stock&', 'q', regex, 'https://www.beiresources.org')
        
