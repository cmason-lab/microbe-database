#!/usr/bin/env python
# coding=utf-8

import openpyxl
import re
from validate import *
from config import *

path = 'annotations_master 12-19-16 - Copy.xlsx'
wb = openpyxl.load_workbook(filename=path)
ws = wb.get_active_sheet()

for row in range(2, ws.max_row+1):
    for key in keys_to_cols:
        col = keys_to_cols[key]
        type = type_from_key[key]
        
        value = str(ws.cell(row = row, column = col).value)
        
        if value == '#N/A' or value == 'N/A' or value =='None':
            continue
        
        if validate(value, type) is False:
            value = '{} badformat {}'.format(value, key)
            ws.cell(row = row, column = col).value = value

wb.save(path)