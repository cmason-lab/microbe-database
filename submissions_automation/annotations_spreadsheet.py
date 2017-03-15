#!/usr/bin/env python3

import os
import pickle

class AnnotationsSpreadsheet():
    ''' Allows processing of annotation .xlsx files'''

    def __init__(self, spreadsheet_path):
        self.spreadsheet_path = spreadsheet_path   
        # See if we have a pickled object (loads way faster)
        self.pickled_spreadsheet_path = self.spreadsheet_path[0:-5] + '.pickle'
        
        if os.path.isfile(self.pickled_spreadsheet_path):
            print('Loading pickle: ' + self.pickled_spreadsheet_path)
            self.data = pickle.load(open(self.pickled_spreadsheet_path, 'rb'))
            print('Pickle loaded')
        else:
            import openpyxl

            wb = openpyxl.load_workbook(filename=self.spreadsheet_path)
            ws = wb.get_active_sheet()

            self.data = []
            for i in range(1, ws.max_row+1):
                print('Reading spreadsheet row ' + i + '/' + ws.max_row)
                row = []
                for j in range(1, ws.max_column+1):
                    row.append(ws.cell(row=i, column=j).value)
                    
                self.data.append(row)
            
            print('Saving pickle: ' + self.pickled_spreadsheet_path)
            pickle.dump(self.data, open(self.pickled_spreadsheet_path, 'wb'))
            print('Pickle saved.')
        
        # For iteration
        self.current_row = 1
        self.max_row = len(self.data)
        self.max_column = len(self.data[0])
        
    def getCell(self, row, column):
        return self.data[row-1][column-1]
    
    def __iter__(self):
        return self

    def __next__(self):
        if self.current_row > self.max_row:
            self.current_row = 1
            raise StopIteration
        else:
            row = []
            for column in range(0, self.max_column):
                row.append(self.getCell(self.current_row, column))
            self.current_row += 1
            return row