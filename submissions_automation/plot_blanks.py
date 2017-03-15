#!/bin/usr/env python3

import openpyxl
from config import *
import shutil
import math
import os

# How many people are STILL working on this project
num_submitters = 42
# How many annotations does each person do per week?
annotations_per_week = 10

read_wb = openpyxl.load_workbook(filename = MASTER_WS_PATH)
read_ws = read_wb.get_active_sheet()

entries = []

for row in range(2, read_ws.max_row+1):
    full_taxonomy = read_ws.cell(row = row, column = 1).value
    genus_species = read_ws.cell(row = row, column = 2).value
    entries.append({'full_taxonomy': full_taxonomy, 'genus_species': genus_species})

# Splits the num of entries to be completed amongst all submitters
assignments = []
#annotations_per_submitter = math.ceil(len(entries)/num_submitters)

def is_blank_value(value):
    if value == '' or value == None:
        return True
    else:
        return False
    
cells_to_check = list(range(3, 59, 2))
cells_to_check.append(58)

blank_count = {}

for row in range(2, len(entries)+1):
    blank_cols = 0
    for col in cells_to_check:
        cell_value = read_ws.cell(row=row, column=col).value
        if is_blank_value(cell_value):
            blank_cols = blank_cols + 1
    
    if(blank_cols in blank_count):
        blank_count[blank_cols] = blank_count[blank_cols] + 1
    else:
        blank_count[blank_cols] = 1
        

from plotly.offline import plot
import plotly.graph_objs as go

keys = list(blank_count.keys())
values = list(blank_count.values())

data = [go.Bar(
    x=keys,
    y=values
)]

layout = go.Layout(
    xaxis=dict(
        title='Number of Blanks',
        autotick=False
    ),
    yaxis=dict(
        title='Count'
    )
)
fig = go.Figure(data=data, layout=layout)
plot(fig, image='png', image_filename='blank_counts', filename='blank_counts.html', auto_open=False)