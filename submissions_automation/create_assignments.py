#!/bin/usr/env python3

import openpyxl
from config import *
import shutil
import math
import os

# How many people are working on this project
num_submitters = 42
# How many annotations does each person do per week?
annotations_per_week = 10

read_wb = openpyxl.load_workbook(filename = ANN_TO_ASSIGN_PATH)
read_ws = read_wb.get_active_sheet()

entries = []

for row in range(2, read_ws.max_row+1):
    full_taxonomy = read_ws.cell(row = row, column = 1).value
    genus_species = read_ws.cell(row = row, column = 2).value
    entries.append({'full_taxonomy': full_taxonomy, 'genus_species': genus_species})

# Splits the num of entries to be completed amongst all submitters
assignments = []
#annotations_per_submitter = math.ceil(len(entries)/num_submitters)

for i in range(0, len(entries), annotations_per_week):
    #annotations_for_submitter = entries[i:i+annotations_per_submitter]
    annotations_assignment = entries[i:i+annotations_per_week]
    assignments.append(annotations_assignment)

# Create a .xlsx for each assignment
for i in range(0, len(assignments)):
    print('Creating assignment {}'.format(i+1))
    
    dir = '{}{}'.format(UNASSIGNED_ASSIGNMENTS_DIR, i+1)
    # Remove the old dir if it exists
    if os.path.exists(dir):
        shutil.rmtree(dir)
    os.makedirs(dir)
    
    assignment = assignments[i]
    assignment_xlsx = '{}\\{}\\{}.xlsx'.format(UNASSIGNED_ASSIGNMENTS_DIR, i+1, i+1)
    
    # Copy the template (do this to preserve formatting)
    shutil.copyfile(TEMPLATE_PATH, assignment_xlsx)
        
    # Edit the template you just copied
    write_wb = openpyxl.load_workbook(filename = assignment_xlsx)
    write_ws = write_wb.get_active_sheet()
        
    # Start filling things in at row 2
    row = 2
        
    for entry in assignment:
        full_taxonomy = entry['full_taxonomy']
        genus_species = entry['genus_species']
            
        write_ws.cell(row = row, column = 1).value = full_taxonomy
        write_ws.cell(row = row, column = 2).value = genus_species
        row = row + 1

    # Save the changes
    write_wb.save(assignment_xlsx)