#!/usr/bin/env python3

import openpyxl

#read_path = 'summer_2016/presummer_annotations_master.xlsx'
master_path = 'annotations_master.xlsx'

# Start to read annotations from the old master annotations
read_wb = openpyxl.load_workbook(filename = read_path)
read_ws = read_wb.get_active_sheet()

write_wb = openpyxl.load_workbook(filename = master_path)
write_ws = write_wb.get_active_sheet()

write_row_from_organism_name = {}

# You have to iterate through the entire write_ws in order to assign values by organism name
# Stupid, but there's no other way to do this
for row in range(2, 7679):
    # Full taxonomy is in column 1
    cell = write_ws.cell(row = row, column = 1)
    write_row_from_organism_name[cell.value] = row

# Copy cells from read_ws to write_ws
def copy_col( copy_row, copy_col, write_row, write_col ):
    copy_cell = read_ws.cell(row = copy_row, column = copy_col)
    write_cell = write_ws.cell(row = write_row, column = write_col)
    #print(copy_cell.value, write_cell.value)
    write_cell.value = copy_cell.value

# Iterate through read_ws and copy values
for read_row in range(2, read_ws.max_row):
    cell = read_ws.cell(row = read_row, column = 1)
    write_row = write_row_from_organism_name[cell.value]
    
    for i in range(2, 30):
        copy_col(read_row, i, write_row, i*2 - 1)
    
    # No citation for picture URL messes up our pattern (bc picture URL is inherently a citation)
    copy_col(read_row, 31, write_row, 58)

    # Citations
    for i in range(4, 58, 2):
        copy_col(read_row, 30, write_row, i)
        
    copy_col(read_row, 30, write_row, 59)
        
write_wb.save(master_path)