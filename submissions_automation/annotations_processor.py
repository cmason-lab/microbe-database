#!/usr/bin/env python3

import openpyxl

class AnnotationsProcessor():
    ''' Allows processing of annotation .xlsx files

Move and process xlsx files by column and organism name'''

    def __init__(self, read_ws_path, write_ws_path):
        # Hardcode this bc there may be some problems if students make random inserts
        #self.write_max_row = 11
        #self.write_max_column = 59
        
        self.read_ws_path = read_ws_path
        self.write_ws_path = write_ws_path
        
        self.read_wb = openpyxl.load_workbook(filename = self.read_ws_path)
        self.read_ws = self.read_wb.get_active_sheet()
        
        self.write_wb = openpyxl.load_workbook(filename = self.write_ws_path)
        self.write_ws = self.write_wb.get_active_sheet()
        
        # This is a dictionary  that will return the row of a given
        # organism based on the organism's name
        self.write_row_from_organism_name = {}
        
        # You have to iterate through the entire write_ws in order to assign values by organism name
        # Stupid, but there's no other way to do this
        for row in range(2, self.write_ws.max_row+1):
            # Full taxonomy is in column 1
            cell = self.write_ws.cell(row = row, column = 1)
            self.write_row_from_organism_name[cell.value] = row

    def copy_cell(self, copy_row, copy_col, write_row, write_col):
        ''' Copy cells from read_ws to write_ws '''
        
        copy_cell = self.read_ws.cell(row = copy_row, column = copy_col)
        write_cell = self.write_ws.cell(row = write_row, column = write_col)
        write_cell.value = copy_cell.value
        
    def update(self):
        ''' Update the write_ws with data from the read_ws '''
        
        # Iterate through read_ws and copy values
        for read_row in range(2, self.read_ws.max_row+1):
            # Get the name of the organism
            cell = self.read_ws.cell(row = read_row, column = 1)
            # See where this organism is at in the write_ws
            write_row = self.write_row_from_organism_name[cell.value]
            
            # Update each column in the write_ws with that from read_ws
            for i in range(3, self.read_ws.max_column+1):
                self.copy_cell(read_row, i, write_row, i)

    def save(self):
        ''' Save changes to the write_wb '''
        
        self.write_wb.save(self.write_ws_path)