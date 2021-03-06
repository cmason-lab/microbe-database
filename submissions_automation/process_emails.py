#!/usr/bin/env python3

import imaplib
import email
import sys
import os
import openpyxl
import shutil

import time
import datetime

import email_assistant

# Import all configuration variables from config.py; includes logging
from config import *

# ------------------------- CONNECT TO GMAIL

# Must enable imap in gmail settings
mail = imaplib.IMAP4_SSL('imap.gmail.com')
mail.login(USERNAME, PASSWORD)

# ------------------------- SEARCH FOR UNPROCESSED EMAILS

# All emails automatically go to 'Unprocessed' (via gmail filters)
mail.select('Unprocessed')
response, data = mail.search(None, 'ALL')
# Email ids are underscore-delimited and reside at data[0]
unprocessed_email_ids = data[0].split()

# Exit the script if there's nothing to process
if len(unprocessed_email_ids) == 0:
    exit()

# ------------------------- PROCESS EACH EMAIL

# Fxn that gets a python representation of the raw email from the server
def get_email_by_id( email_id ):
    # Get the raw data of the email
    response, data = mail.fetch(email_id, "(RFC822)")
    # The raw message is in the data arrray at [0][1], not sure why... 
    return email.message_from_bytes(data[0][1])

def change_email_label( email_id, label ):
    # Add the new label
    mail.copy(email_id, label)
    # Delete the old label
    mail.store(email_id , '+FLAGS', '(\Deleted)')
    mail.expunge()

# You have to do reversed, bc after you process an email, you change the IDs
# of all emails on the server. Going in reverse will prevent changing index problems
for email_id in reversed(unprocessed_email_ids):
    email_msg = get_email_by_id(email_id)
    
    # Parse the emails by subject to see what was submitted
    subject = email_msg['Subject']
    submitter_email = email_msg['From']
    
    # Gmail automatically adds sent emails to the inbox...
    if(subject == 'Mason Lab Microbe Database - New Assignment'):
        print('From us, skipping email.')
        continue
    
    # Subject line is formatted as {Submitter ID} {Assignment ID}
    # Try to parse this. If there's an error, write this in the log
    try:
        subject_split = subject.split('_')
        submitter_id = subject_split[0];
        assignment_id = subject_split[1];
    except:
        # Record the error
        log('SUBJECT of email {} is formatted incorrectly.'.format(email_msg['Message-ID']))
        # Move the email from 'Unprocessed' to 'Error'
        change_email_label(email_id, 'Error')
        
        # Skip further processing of this email
        continue
    
    submit_date = email_msg['Date']
    
    has_attachment = False
    is_update = False
    
    # Sloppy way to break out of email outer loop
    break_outer_email_loop = False
    
    # Pull out and save the attachment
    if email_msg.is_multipart():
        for part in email_msg.walk():
            if part.get_content_maintype() == 'multipart':
                continue
            if part.get('Content-Disposition') is None:
                continue
            
            filename = part.get_filename()
            extension = os.path.splitext(filename)[1]
            
            if filename is not None:
                # This is where we will save/download the attached submission; same as subject but w/ .xlsx extension
                submission_save_path = os.path.join(UNPROCESSED_SUBMISSIONS_DIR, '{}.xlsx'.format(subject))
                # This is where the file would reside if it had already been submitted and processed (used to see if they are submitting an update)
                processed_save_path = os.path.join(PROCESSED_SUBMISSIONS_DIR, '{}.xlsx'.format(subject))
                
                # Make sure we've actually grabbed the .xlsx from the email
                if extension == '.xlsx':
                    # See if the submission already exists (making this an update), but it hasn't been processed yet
                    if os.path.isfile(submission_save_path) and extension == '.xlsx':
                        print('Updating the file at {}'.format(submission_save_path))
                        # Easy peasy just update the file in the unprocessed dir
                        change_email_label(email_id, 'Update')
                        # Delete the old assignment and replace it with the new one
                        os.remove(submission_save_path)
                        # Save the attachment in this same spot
                        with open(submission_save_path, 'wb') as f:
                            f.write(part.get_payload(decode=True))
                            
                        # Annnd we're done (break out of the outer loop--the email loop)
                        break_outer_email_loop = True
                        continue
                    elif os.path.isfile(processed_save_path):
                        print('This is an update for the the already processed file at {}'.format(processed_save_path))
                        change_email_label(email_id, 'Update')
                        
                        has_attachment = True
                        is_update = True
                        
                        # Save the attachment in the unprocessed dir--it will overwrite the master later
                        with open(submission_save_path, 'wb') as f:
                            f.write(part.get_payload(decode=True))
                            has_attachment = True
                    else:
                        # All clear--save the attachment in the unprocessed dir
                        with open(submission_save_path, 'wb') as f:
                            f.write(part.get_payload(decode=True))
                            has_attachment = True
                else:
                    log('Error saving attachment, {}, for email {}. File is not a .xlsx file'.format(submission_save_path, email_msg['Message-ID']))
                    # Move the email from 'Unprocessed' to 'Error'
                    change_email_label(email_id, 'Error')
                    break_outer_email_loop = True
    
    if break_outer_email_loop == True:
        print('continuing with email loop')
        continue
    
    # Log an error if we can't find an attachment (this might not get picked up by our if then tree above)
    if has_attachment is not True:
        log('Error with attachment for email {}.'.format(email_msg['Message-ID']))
        # Move the email from 'Unprocessed' to 'Error'
        change_email_label(email_id, 'Error')
        continue
    
    if is_update == False:
        # Move the email from 'Unprocessed' to 'Processed'
        change_email_label(email_id, 'Processed')
        
    # Acknowledge the submission
    wb = openpyxl.load_workbook(filename = SUBMITTERS_PATH)
    submitter_ws = wb['{}'.format(submitter_id)]
    
    # Create a reverse lookup dictionary
    write_row_from_annotation_id = {}
    
    for row in range(2, submitter_ws.max_row+1):
        cell = submitter_ws.cell(row = row, column = 1)
        write_row_from_annotation_id[cell.value] = row
        
    write_row = write_row_from_annotation_id[assignment_id]
    
    if is_update == True:
        print('is update')
        # Use col 4 for updates
        submitter_ws.cell(row = write_row, column = 4).value = submit_date
        wb.save(SUBMITTERS_PATH)
        
        # No further processing (i.e. sending out a new assignment)
        continue
    else:
        submitter_ws.cell(row = write_row, column = 3).value = submit_date
    
    # Sloppy... Chooses which new file to send out
    do_break = False
    # Send out their new assignement
    for subdir, dirs, files in os.walk(UNASSIGNED_ASSIGNMENTS_DIR):
        if do_break == True:
            break
        
        for file in files:
            file_path = os.path.join(subdir, file)
            extension = os.path.splitext(os.path.basename(file_path))[1]
            
            if extension == '.xlsx':
                print(file_path)
                new_assignment_id = os.path.splitext(os.path.basename(file_path))[0]
                do_break = True
                break
   
    # Now that we've decided what assignment to give, actually give it
    if(new_assignment_id):
        save_row = submitter_ws.max_row+1
        
        # Move the old assignment
        shutil.move('{}{}/'.format(UNASSIGNED_ASSIGNMENTS_DIR, new_assignment_id), '{}{}/'.format(ASSIGNED_ASSIGNMENTS_DIR, new_assignment_id))
                    
        # Attach everything in this dir
        attachment_dir = '{}{}'.format(ASSIGNED_ASSIGNMENTS_DIR, new_assignment_id)
        attachments = ['{}{}/{}'.format(ASSIGNED_ASSIGNMENTS_DIR, new_assignment_id, file) for file in os.listdir(attachment_dir)]

        asst = email_assistant.EmailAssistant(USERNAME, PASSWORD)
        asst.send_email(submitter_email, 'Mason Lab Microbe Database - New Assignment', 'Thank you for your previous submission! This set of annotations corrresponds to assignment ID {}. Recall, your submitter ID is {}. Whenever you submit this assignment, the subject line must read {}_{}. This is extremely important!! Please do not reply to this email with anything but submissions! Send all questions to hebashaaban1@gmail.com and daw2035@med.cornell.edu. Keep up the good work!'.format(new_assignment_id, submitter_id, submitter_id, new_assignment_id), attachments)
        
        # Log that we sent out a new assignment
        submitter_ws.cell(row = save_row, column = 1).value = new_assignment_id
        
        # Timestamp
        ts = time.time()
        timestamp = datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')
        submitter_ws.cell(row = save_row, column = 2).value = timestamp
        wb.save(SUBMITTERS_PATH)
    