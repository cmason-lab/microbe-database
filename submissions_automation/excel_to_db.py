#!/usr/bin/env python3
# coding=utf-8

import openpyxl
from config import *
from validate import *
import os
import sqlite3
import re

# Going to compile optimal pH (13), optimal temperature (15), pathogenicity (19),
# sens to antibiotics (29), spore-forming (49), biofilm formingz (51), is extremophile (3),
# Gram-stain (41), found in microbiome (7)

path = 'annotations_master 12-19-16.xlsx'
wb = openpyxl.load_workbook(filename = path)
ws = wb.get_active_sheet()

# -------------
# Make the db in python
# -------------

python_db = {}

for row in range(2, ws.max_row+1):
    species = str(ws.cell(row = row, column = 2).value)
    python_db[species] = {}
    
    # Add the kingdom, phylum, ... stuff
    taxa = str(ws.cell(row = row, column = 1).value)
    expr = re.compile('[a-z]__([A-Za-z]+)')
    while expr.match(taxa):
        taxa = re.sub(expr, '\\1', taxa)
    taxa = taxa.replace('_', ' ')
    taxa = taxa.split('|')
    taxa_levels = ['kingdom', 'phylum', 'class', 'order', 'family', 'genus', 'species']
    
    i = 0
    for taxa_level in taxa_levels:
        python_db[species][taxa_level] = taxa[i]
        i = i+1

    # Add the other values from the spreadsheet
    for key in keys_to_cols:
        value = str(ws.cell(row = row, column = keys_to_cols[key]).value)
        type = type_from_key[key]
        
        if validate(value, type):
            python_db[species][key] = value

# -------------
# Make the db in sql
# -------------

db_path = 'microbe.db'

# Delete the old db (if it exists) and remake it
if os.path.isfile(db_path):
    os.remove(db_path)
    
connection = sqlite3.connect(db_path)
cursor = connection.cursor()
    
# Create tables
tables_sql = {'Microbe': '', 'Antimicrobial': '', 'MicrobiomeLocation': '', 'ExtremeEnvironment': '', 'Citation': ''}

tables_sql['Microbe'] = '''CREATE TABLE Microbe (
    'microbe_id' INTEGER PRIMARY KEY AUTOINCREMENT,
    'kingdom' TEXT NOT NULL,
    'phylum' TEXT NOT NULL,
    'class' TEXT NOT NULL,
    'order' TEXT NOT NULL,
    'family' TEXT NOT NULL,
    'genus' TEXT NOT NULL,
    'species' TEXT NOT NULL,
    'optimal_pH' TEXT,
    'optimal_temperature' TEXT,
    'pathogenicity' INTEGER,
    'antimicrobial_susceptibility' INTEGER,
    'spore_forming' INTEGER,
    'biofilm_forming' INTEGER,
    'extreme_environment' INTEGER,
    'microbiome_location' INTEGER,
    'plant_pathogen' INTEGER,
    'animal_pathogen' INTEGER,
    'gram_stain' INTEGER
);'''

tables_sql['Antimicrobial'] = '''CREATE TABLE Antimicrobial (
    'antimicrobial_id' INTEGER PRIMARY KEY AUTOINCREMENT,
    'name' TEXT NOT NULL
);'''

tables_sql['MicrobiomeLocation'] = '''CREATE TABLE MicrobiomeLocation (
    'microbiome_location_id' INTEGER PRIMARY KEY AUTOINCREMENT,
    'name' TEXT NOT NULL
);'''

tables_sql['ExtremeEnvironment'] = '''CREATE TABLE ExtremeEnvironment (
    'sample_id' INTEGER PRIMARY KEY,
    'name' TEXT NOT NULL
);'''

tables_sql['Citation'] = '''CREATE TABLE Citation (
    'citation_id' INTEGER PRIMARY KEY,
    'microbe_id' INTEGER NOT NULL,
    'microbe_key' TEXT NOT NULL,
    'key_indices' TEXT NOT NULL DEFAULT '0',
    'citation' TEXT NOT NULL,
    'access_date_timestamp' INTEGER NOT NULL,
    FOREIGN KEY('microbe_id') REFERENCES Microbe('microbe_id')
);'''

for table in tables_sql:
    cursor.execute(tables_sql[table])

# -------------
# Add the python data to the db
# -------------

def add_quotes(list):
    new_list = []
    for item in list:
        item = "'{}'".format(item)
        new_list.append(item)
    return new_list

for species in python_db:
    keys = python_db[species].keys()
    values = python_db[species].values()
    
    if(len(values) == 0):
        continue
    
    keys = add_quotes(keys)
    values = add_quotes(values)
    
    keys = ','.join(keys)
    values = ','.join(values)
    
    cursor.execute("INSERT INTO 'Microbe' ({}) VALUES ({})".format(keys, values))

connection.commit()
connection.close()