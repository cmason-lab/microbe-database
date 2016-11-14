#!/bin/usr/env python3

import openpyxl
from config import *

import shutil

read_wb = openpyxl.load_workbook(filename = MASTER_WS_PATH)
read_ws = read_wb.get_active_sheet()

empty_entries = []

for row in range(2, read_ws.max_row+1):
    cell_is_empty = True
    
    # Loop through the col's--if you come across something that is filled in then don't add this
    for col in range(3, 59+1):
        cell = read_ws.cell(row = row, column = col)
        val = cell.value
        
        if not (cell.value == '#N/A' or cell.value == None or cell.value == 0 or cell.value == '0'):
            # Cell is not empty
            cell_is_empty = False
            break
    
    # If still empty, add it
    if cell_is_empty:
        full_taxonomy = read_ws.cell(row = row, column = 1).value
        genus_species = read_ws.cell(row = row, column = 2).value
        empty_entries.append({'full_taxonomy': full_taxonomy, 'genus_species': genus_species})
    
# Copy the template (do this to preserve formatting)
shutil.copyfile(TEMPLATE_PATH, ANN_TO_ASSIGN_PATH)

# Edit the template you just copied
write_wb = openpyxl.load_workbook(filename = ANN_TO_ASSIGN_PATH)
write_ws = write_wb.get_active_sheet()

# Start filling things in at row 2
row = 2

for empty_entry in empty_entries:
    full_taxonomy = empty_entry['full_taxonomy']
    genus_species = empty_entry['genus_species']
    
    write_ws.cell(row = row, column = 1).value = full_taxonomy
    write_ws.cell(row = row, column = 2).value = genus_species
    row = row + 1

# Save the changes
write_wb.save(ANN_TO_ASSIGN_PATH)