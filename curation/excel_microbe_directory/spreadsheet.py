import openpyxl
import sqlite3
import os
import re

from . import cell_validator_factory
from . import spreadsheet_config

from excel_microbe_directory.spreadsheet_config import Columns
from excel_microbe_directory.spreadsheet_config import Taxa
from excel_microbe_directory.spreadsheet_config import ColumnValidationTypes
from excel_microbe_directory.spreadsheet_config import SqlDataTypes

class Spreadsheet():
    def __init__(self, spreadsheet_path):
        self.config = spreadsheet_config.SpreadsheetConfig()
        self.spreadsheet_path = spreadsheet_path
        
        wb = openpyxl.load_workbook(filename = spreadsheet_path)
        self.ws = wb.get_active_sheet()
    
    def export_as_dictionary(self):
        '''openpyxl worksheet --> python dictionary
        
        Access via: dict[Genus species][column name]'''
        try: return self.dict
        except: dict = {}
        
        for row in range(2, self.ws.max_row+1):
            species = self.ws.cell(row = row, column = 2).value
            dict[species] = {}
            
            # Split the long taxonomy string (column 1) into discrete taxa and use these as keys
            taxa = self.__split_taxa_string(self.ws.cell(row = row, column = 1).value)
            for i,taxon in enumerate(self.config.getTaxaHierarchy()):
                dict[species][taxon] = taxa[i]
        
            # Add the other columns from the spreadsheet
            for column_name in self.config.getColumnNames():
                column_number = self.config.getColumnNumber(column_name)
                column_validation_type = self.config.getColumnValidationType(column_name)
                cell_value = str(self.ws.cell(row=row, column=column_number))
                
                if self.validate_cell(cell_value, column_validation_type):
                    dict[species][key] = cell_value
        
        self.dict = dict
        return dict
    
    def validate_cell(self, value, type):
        ''' Ensures that the cells of each column conform to the column's type (as defined by config)'''
        
        validator_factory = cell_validator_factory.MicrobeDirectoryCellValidatorFactory()
        
        if type == ColumnValidationTypes.COGEM_PATHOGENICITY:
            validator = validator_factory.createCogemPathogenicityValidator()
        elif type == ColumnValidationTypes.RANGE:
            validator = validator_factory.createRangeValidator()
        elif type == ColumnValidationTypes.BINARY:
            validator = validator_factory.createBinaryValidator()
        elif type == ColumnValidationTypes.TERNARY:
            validator = validator_factory.createTernaryValidator()
        
        return validator.isValid(value)
    
    def export_as_sqlite(self, db_path):
        try: self.dict
        except: self.export_as_dictionary()
        
        # Delete the old db (if it exists) and remake it
        if os.path.isfile(db_path):
            os.remove(db_path)
            
        connection = sqlite3.connect(db_path)
        cursor = connection.cursor()
            
        # -------------
        # Create the tables
        # -------------
        tables_sql = {}

        taxa_sql = ''
        for taxon in self.config.getTaxaHierarchy():
            taxa_sql += "'{taxon}' TEXT,".format(taxon=taxon)
        taxa_sql = taxa_sql[:-1]
        
        columns_sql = ''
        for column_name in self.config.getColumnNames():
            columns_sql += "'{column_name}' {column_type},".format(column_name=column_name, column_type=self.config.getSqlDataType(column_name))
        columns_sql = columns_sql[:-1]
        
        tables_sql['Microbe'] = '''CREATE TABLE Microbe (
            'microbe_id' INTEGER PRIMARY KEY AUTOINCREMENT,
            {taxa_sql},
            {columns_sql}
        );'''.format(taxa_sql=taxa_sql, columns_sql=columns_sql)
        
        tables_sql['Antimicrobial'] = '''CREATE TABLE Antimicrobial (
            'antimicrobial_id' INTEGER PRIMARY KEY AUTOINCREMENT,
            'name' TEXT NOT NULL
        );'''
        
        tables_sql['MicrobiomeLocation'] = '''CREATE TABLE MicrobiomeLocation (
            'microbiome_location_id' INTEGER PRIMARY KEY AUTOINCREMENT,
            'name' TEXT NOT NULL
        );'''
        
        tables_sql['ExtremeEnvironment'] = '''CREATE TABLE ExtremeEnvironment (
            'sample_id' INTEGER PRIMARY KEY,
            'name' TEXT NOT NULL
        );'''
        
        tables_sql['Citation'] = '''CREATE TABLE Citation (
            'citation_id' INTEGER PRIMARY KEY,
            'microbe_id' INTEGER NOT NULL,
            'microbe_key' TEXT NOT NULL,
            'key_indices' TEXT NOT NULL DEFAULT '0',
            'citation' TEXT NOT NULL,
            'access_date_timestamp' INTEGER NOT NULL,
            FOREIGN KEY('microbe_id') REFERENCES Microbe('microbe_id')
        );'''
        
        for table in tables_sql:
            cursor.execute(tables_sql[table])
        
        # -------------
        # Populate the tables
        # -------------
        for species in self.dict:
            keys = self.dict[species].keys()
            values = self.dict[species].values()
            
            if(len(values) == 0):
                continue
            
            keys = ["'{}'".format(key) for key in keys]
            values = ["'{}'".format(value) for value in values]
            
            keys = ','.join(keys)
            values = ','.join(values)
            
            cursor.execute("INSERT INTO 'Microbe' ({}) VALUES ({})".format(keys.lower(), values))
        
        connection.commit()
        connection.close()
        
    def __split_taxa_string(self, taxa):
        '''Splits values from column 1 of spreadsheet into a list of discrete taxa'''
        expr = re.compile('[a-z]__([A-Za-z]+)')
        while expr.match(taxa):
            taxa = re.sub(expr, '\\1', taxa)
        taxa = taxa.replace('_', ' ')
        taxa = taxa.split('|')
        return taxa